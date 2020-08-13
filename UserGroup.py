import pyad
import pyad.adquery
import win32api
from pyad import *

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'User_Group.xlsx'
ws1 = wb.active
ws1.title = "User_Grup"

listOU =['OU=09 - Usuarios Aguas,OU=01 - Aguas de Guariroba,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=03 - Aguas de Campo Verde,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=04 - Aguas de Carlinda,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=06 - Aguas de Claudia,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=07 - Aguas de Jauru,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=08 - Aguas de Marcelandia,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=09 - Aguas de Nortelandia,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=10 - Aguas de Pocone,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=11 - Aguas de Porto Esperidiao,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=12 - Aguas de Primavera,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=13 - Aguas de Santa Carmem,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=14 - Aguas de Sao Jose,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=15 - Aguas de Sorriso,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=16 - Aguas de Uniao do Sul,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=17 - Aguas de Vera,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=18 - Aguas de Peixoto,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=19 - Saneamento de Jangada,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=20 - Saneamento Pedra Preta,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=21 - Aguas de Confresa,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=21 - Diamantino,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=22 - Consorcio Ibura,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=23 - Aguas de Guaranta,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=24 - Aguas de Novo Progresso,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=25 - Aguas de Matupa,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=26 - Sinop,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=27 - Aguas de Paranatinga,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=28 - Aguas de Buritis,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=29 - Aguas de Pimenta Bueno,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=30 - Cuiaba,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=31 - Aguas de Rolim de Moura,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=32 - Aguas de Ariquemes,OU=02 - Usuarios,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=01 - Sarpav,OU=04 - Mineradoras,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=02 - Minerpav,OU=04 - Mineradoras,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=05 - Usuarios,OU=06 - Prolagos,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=05 - Usuarios,OU=07 - Univias,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuários,OU=08 - Aguas de Manaus,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=10 - Aegea,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=1.3 - Contas que não expiram,OU=01 - Usuarios,OU=10 - Aegea,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=1.5 - Diretoria,OU=01 - Usuarios,OU=10 - Aegea,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuários,OU=11 - Aguas de Barra do Garca,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=14 - Equipav Engenharia,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=1.1 - Usuarios Obras,OU=01 - Usuarios,OU=14 - Equipav Engenharia,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=15 - Aguas do Mirante,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=17 - CAA,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuários,OU=18 - Aguas de Sao Francisco,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuários,OU=19 - LVEBR,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=20 - Aguas de Matao,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=21 - GSS,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=22 - Sao Francisco do Sul,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=24 - Aguas de Timon,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=25 - Aegea Engenharia,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=26 - Aguas de Penha,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=27 - Aguas de Holambra,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=28 - Aguas de Camboriu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=29 - Serra Ambiental,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=30 - Aguas de Bombinhas,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=31 - Aguas de Teresina,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=32 - Vila Velha Ambiental,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net',
        'OU=01 - Usuarios,OU=33 - Ambiental Metrosul,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net']

cont = 0
valid = True

row = 0
col = 1

pyad.set_defaults(ldap_server="", username="", password="")
#ou = pyad.adcontainer.ADContainer.from_dn("OU=01 - Usuarios,OU=21 - GSS,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net")

for ouUser in listOU:
    ou = pyad.adcontainer.ADContainer.from_dn(ouUser)
    print('---------OU---------')
    print(ouUser)
    for comp in ou.get_children():
        userName = comp.get_attribute('CPF')
        col = 1
        if len(userName) > 0:
            row += 1
            _=ws1.cell(column=row, row=col, value=userName[0])
            print ("----------------")
            print (userName)
            grupMemberOf = comp.get_attribute('MemberOf')
            valid = True
            while valid == True:
                if  cont < len(grupMemberOf):
                    try:
                        ouGroup = pyad.adcontainer.ADContainer.from_dn(grupMemberOf[cont])
                    except:
                        print('Erro')
                    namGroup = ouGroup.get_attribute('cn')
                    
                    col += 1
                    _=ws1.cell(column=row, row=col, value = namGroup[0])
                    
                    print (namGroup)
                    cont = 1 + cont
                else:
                    cont = 0
                    valid = False
wb.save(filename = dest_filename)
