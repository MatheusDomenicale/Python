from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

list1 = ['matheus', 'Adriana', 'Hermeson', 'Tonhao', 'Fabricio']
list2 = ['grupo 1', 'grupo 2', 'grupo 3']
row = 0
col = 1

for x in list1:
    row += 1
    col = 1
    _= ws1.cell(column=row, row=col, value=x)
    for y in list2:
        col += 1
        _= ws1.cell(column=row, row=col, value=y)

wb.save(filename = dest_filename)












