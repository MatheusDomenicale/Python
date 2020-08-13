import pyad
import pyad.adquery
import win32api
from pyad import *

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = "User_Grup"


cont = 0
valid = True

row = 0
col = 1

pyad.set_defaults(ldap_server="", username="", password="")
ou = pyad.adcontainer.ADContainer.from_dn("OU=01 - Usuarios,OU=21 - GSS,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net")

for comp in ou.get_children():
    ntbesp = comp.get_attribute('cn')
    col = 1
    if len(ntbesp) > 0:
        row += 1
        _=ws1.cell(column=row, row=col, value=ntbesp[0])
        print ("----------------")
        print (ntbesp)
        grupMemberOf = comp.get_attribute('MemberOf')
        valid = True
        while valid == True:
            if  cont < len(grupMemberOf):
                ouGroup = pyad.adcontainer.ADContainer.from_dn(grupMemberOf[cont])
                namGroup = ouGroup.get_attribute('cn')
                
                col += 1
                _=ws1.cell(column=row, row=col, value = namGroup[0])
                
                print (namGroup)
                cont = 1 + cont
            else:
                cont = 0
                valid = False
wb.save(filename = dest_filename)
