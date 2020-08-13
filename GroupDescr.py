import pyad
import pyad.adquery
import win32api
from pyad import *

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'User_Group.xlsx'
ws1 = wb.active
ws1.title = "User_Grup"

listOU =['OU=02 - Grupos Distribuicao,OU=01 - Aguas de Guariroba,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=03 - Grupos Globais,OU=01 - Aguas de Guariroba,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=08 - Grupos Antigos - Domain Aguas,OU=01 - Aguas de Guariroba,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=14 - Grupos Domain Local - M,OU=01 - Aguas de Guariroba,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=14 - Grupos Domain Local - R,OU=01 - Aguas de Guariroba,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=18 - Scanner e Impressoras,OU=01 - Aguas de Guariroba,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=03 - Grupos Domain Local - M,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=04 - Grupos Domain Local - R,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Globais,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupos de Distribuicao,OU=02 - Nascentes do Xingu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=02 - Grupos Distribuicao,OU=03 - Cibe,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=02 - Grupos Domain Local,OU=01 - Sarpav,OU=04 - Mineradoras,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=03 - Grupos Domain Local,OU=02 - Minerpav,OU=04 - Mineradoras,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=02 - Grupos de Distribuicao,OU=06 - Prolagos,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=03 - Grupos Globais,OU=06 - Prolagos,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=03 - Grupos Globais,OU=07 - Univias,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=09 - Grupos Domain Local M,OU=07 - Univias,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=10 - Grupos Domain Local R,OU=07 - Univias,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=04 - Grupos de Distribuição,OU=08 - Aguas de Manaus,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos Globais,OU=08 - Aguas de Manaus,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Domain Local - M,OU=08 - Aguas de Manaus,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupos Domain Local - R,OU=08 - Aguas de Manaus,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=04 - Grupos de Distribuição,OU=10 - Aegea,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos Globais,OU=10 - Aegea,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Domain Local - M,OU=10 - Aegea,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupos Domain Local - R,OU=10 - Aegea,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=10 - Grupos EPM,OU=10 - Aegea,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=02 - EPM,OU=11 - Grupos BI-Planning,OU=10 - Aegea,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Globais,OU=12 - Aguas de Meriti,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupo Distribuição,OU=12 - Aguas de Meriti,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=02 - Grupos Antigos Netpav,OU=14 - Equipav Engenharia,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=03 - Grupos Globais,OU=14 - Equipav Engenharia,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=04 - Grupos de Distribuição,OU=15 - Aguas do Mirante,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos Globais,OU=15 - Aguas do Mirante,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Domain Local - M,OU=15 - Aguas do Mirante,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupos Domain Local - R,OU=15 - Aguas do Mirante,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=04 - Grupos de Distribuição,OU=17 - CAA,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos Globais,OU=17 - CAA,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Domain Local - M,OU=17 - CAA,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupos Domain Local - R,OU=17 - CAA,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos Domain Local - M,OU=18 - Aguas de Sao Francisco,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Domain Local - R,OU=18 - Aguas de Sao Francisco,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupos Globais,OU=18 - Aguas de Sao Francisco,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=08 - Grupos de Distribuição,OU=18 - Aguas de Sao Francisco,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos Domain Local - R,OU=20 - Aguas de Matao,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Domain Local - M,OU=20 - Aguas de Matao,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=08 - Grupos Globais,OU=20 - Aguas de Matao,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=09 - Grupos de Distribuicao,OU=20 - Aguas de Matao,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=09 - Grupos de Distribuicao,OU=20 - Aguas de Matao,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=04 - Grupos,OU=21 - GSS,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Globais,OU=21 - GSS,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=04 - Grupos de Distribuição,OU=22 - Sao Francisco do Sul,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos Globais,OU=22 - Sao Francisco do Sul,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Domain Local - M,OU=22 - Sao Francisco do Sul,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupos Domain Local - R,OU=22 - Sao Francisco do Sul,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupo de Distribuição,OU=24 - Aguas de Timon,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Globais,OU=24 - Aguas de Timon,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupos Domain Local - M,OU=24 - Aguas de Timon,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=08 - Grupos Domain Local - R,OU=24 - Aguas de Timon,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupo de Distribuição,OU=25 - Aegea Engenharia,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupos Domain Local - M,OU=25 - Aegea Engenharia,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=08 - Grupos Domain Local - R,OU=25 - Aegea Engenharia,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos de Distribuicao,OU=26 - Aguas de Penha,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Globais,OU=26 - Aguas de Penha,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos de Distribuicao,OU=27 - Aguas de Holambra,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Globais,OU=27 - Aguas de Holambra,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupos Domain Local - M,OU=27 - Aguas de Holambra,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=08 - Grupos Domain Local - R,OU=27 - Aguas de Holambra,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos de Distribuicao,OU=28 - Aguas de Camboriu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Globais,OU=28 - Aguas de Camboriu,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos de Distribuiçao,OU=29 - Serra Ambiental,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Globais,OU=29 - Serra Ambiental,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupos Domain Local - M,OU=29 - Serra Ambiental,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=08 - Grupos Domain Local - R,OU=29 - Serra Ambiental,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos de Distribuicao,OU=30 - Aguas de Bombinhas,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Globais,OU=30 - Aguas de Bombinhas,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos de Distribuição,OU=31 - Aguas de Teresina,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Globais,OU=31 - Aguas de Teresina,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupos Domain Local - M,OU=31 - Aguas de Teresina,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=08 - Grupo Domain Local - R,OU=31 - Aguas de Teresina,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=05 - Grupos de Distribuição,OU=32 - Vila Velha Ambiental,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=06 - Grupos Globais,OU=32 - Vila Velha Ambiental,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=07 - Grupos Domain Local - M,OU=32 - Vila Velha Ambiental,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=08 - Grupos Domain Local - R,OU=32 - Vila Velha Ambiental,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net'
        ,'OU=04 - Grupos Administrativos,OU=Datacenter Global Crossing,DC=latam,DC=corp,DC=net']

cont = 0
row = 0
erros = 0
pyad.set_defaults(ldap_server="", username="", password="")
#ou = pyad.adcontainer.ADContainer.from_dn("OU=01 - Usuarios,OU=21 - GSS,OU=Estrutura Organizacional CIBE,DC=latam,DC=corp,DC=net")

for ouGroup in listOU:
    ou = pyad.adcontainer.ADContainer.from_dn(ouGroup)
    print('---------OU---------')
    print(ouGroup)
    for comp in ou.get_children():
        col = 1
        print('---------Group---------')
        group = comp.get_attribute('cn')
        row += 1
        try:
            _=ws1.cell(column=col, row=row, value=group[0])
        except:
            erros +=1
            print(erros)
        print(group)
        descri = comp.get_attribute('description')
        if len(descri) > 0:
            col = 2
            try:
                _=ws1.cell(column=col, row=row, value=descri[0])
            except:
                erros +=1
                print(erros)
            print('---------Describe---------')
            print(descri)
wb.save(filename = dest_filename)
